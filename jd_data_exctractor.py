import requests 
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup 
import json
import time
import xlsxwriter
start_time = time.time()

number_ref = json.loads("""{ "9d001":"0", "9d002":"1", "9d003":"2", "9d004":"3", "9d005":"4", "9d006":"5", "9d007":"6", "9d008":"7", "9d009":"8", "9d010":"9", "9d011":"+", "":"-", "9d013":")", "9d014":"("}""")
array = []
title_array = []
city_array = []
number_array = []
rating_array = []
address_array = []
website_array = []
agent = {"User-Agent":'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36'}

link_excel=load_workbook('scrap_links.xlsx')
sheet = link_excel.active

for i in range(1,len(sheet['B'])+1):
	URL = sheet['B'+str(i)].value
	print(URL)
	city = sheet['A'+str(i)].value
	r = requests.get(URL, headers=agent)
	soup = BeautifulSoup(r.content, 'html5lib')
	#(1)select css block from header and make an json for icon=>number link
	style = soup.findAll('style')
	number_ref_temp_list = str(style[1]).split('}')
	number_ref_list = {}
	#(1)iterate loop for icon list
	for i in range(2,16):
		temp = number_ref_temp_list[i].split(':');
		number_ref_list[temp[0][1:]] = number_ref[temp[2][2:-1]]#number_ref assign value for number like 9d001 to 0
	#(2)contact box with detail like number, address, website url, ets.
	content = soup.find('ul',attrs = {'class':'comp-contact'})
	#(2)(1)mobile number list
	try:
		number_icon_list = str(content.find('span',attrs={'class':'telnowpr'})).split(',')
		mobile_number = []
		#(2)(1)iterate loop for mobile number generation
		for icon_list in number_icon_list:
			temp = icon_list.split('mobilesv ')
			number_temp = ''
			for te in temp:
				if te[0:4] == 'icon':
					number_temp+=number_ref_list[te.split('"')[0]]
			mobile_number.append(number_temp)
	except:
		print('hffg')
	try:
		website_domain = content.findAll('span',attrs={'class':'mreinfp comp-text'})[1].a['href']
	except:
		website_domain = ''
	try:
		ratings = soup.find('span',attrs={'class':'value-titles'}).text
	except:
		ratings = ''
	try:
		address = content.find('span',attrs={'class':'lng_add'}).text
	except:
		address = ''
	try:
		title = soup.find('span',attrs={'class':'fn'}).text
	except:
		title = ''
	title_array.append(title)
	number_array.append("||".join(str(x) for x in mobile_number))
	rating_array.append(ratings)
	address_array.append(address)
	website_array.append(website_domain)
	city_array.append(city)
	print(mobile_number)
	print(address)
	print(website_domain)
	print(ratings)
	print(title)
	print(city)

array.append(title_array)
array.append(number_array)
array.append(rating_array)
array.append(city_array)
array.append(address_array)
array.append(website_array)

workbook = xlsxwriter.Workbook('data.xlsx')
worksheet = workbook.add_worksheet()
row = 0
for col, data in enumerate(array):
    worksheet.write_column(row, col, data)
workbook.close()
print (array)
